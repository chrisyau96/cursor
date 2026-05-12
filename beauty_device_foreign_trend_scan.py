#!/usr/bin/env python3
"""
Hong Kong sourcing trial: surface *new* foreign-market home beauty-device signals
from XHS (Xiaohongshu), Instagram, TikTok, and X (Twitter) via Apify Actors, then
structure and sanity-check findings with an LLM (Mistral by default).

Design goals
------------
- **3–5 items per social source** (trial-sized pulls), configurable.
- **Dates**: never trust platform metadata alone. The model must reconcile
  `posted_at_hint` with post text (e.g. "2025 新品", "launching next month")
  and emit an ISO date *or* explicit uncertainty.
- **Distributors**: optional second Apify run using the official
  `apify/google-search-scraper` (or your own actor id) with AI-generated
  B2B-oriented queries; results are flattened to candidate URLs.

Secrets
-------
Set **only** via environment variables (never commit tokens):

- `APIFY_API_TOKEN` — required for Apify runs.
- `MISTRAL_API_KEY` — required for JSON extraction unless you wire another client.

Actor configuration
-------------------
Each social network uses a *different* Apify actor and input schema. This script
expects you to choose Store actors and paste ids:

- `APIFY_ACTOR_XHS`          (example format: `username~actor-name`)
- `APIFY_ACTOR_INSTAGRAM`
- `APIFY_ACTOR_TIKTOK`
- `APIFY_ACTOR_X`

Optional overrides (merged on top of built-in defaults):

- `APIFY_INPUT_OVERRIDE_JSON` — JSON object keyed by `xhs`, `instagram`, `tiktok`, `x`.

Distributor Google search actor:

- `APIFY_ACTOR_GOOGLE_SEARCH` (default `apify~google-search-scraper`)

Tuning:

- `MAX_ITEMS_PER_SOURCE` — integer 3–5 (default 4).
- `DAYS_LOOKBACK` — default 90.
- `OUTPUT_XLSX` — output path.
- `MISTRAL_MODEL` — default `mistral-small-latest`.

Run:  `python beauty_device_foreign_trend_scan.py`
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import logging
import os
import re
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, UnidentifiedImageError


# ---------------------------------------------------------------------------
# Configuration (env-first; no hard-coded API keys)
# ---------------------------------------------------------------------------

DEFAULT_SEARCH_TERMS: List[str] = [
    "LED face mask new",
    "home RF beauty device",
    "microcurrent facial device launch",
    "IPL home device 2025",
    "K-beauty device review",
    "美顔器 新作 家電",
    "美容儀 新品 推薦",
    "뷰티 디바이스 신제품",
    "家用美容仪 新品",
]

REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT_SECONDS", "45"))
IGNORE_SYSTEM_PROXY = os.getenv("IGNORE_SYSTEM_PROXY", "true").lower() in (
    "1",
    "true",
    "yes",
)


def _env_int(name: str, default: int, lo: int, hi: int) -> int:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        v = int(raw)
    except ValueError:
        return default
    return max(lo, min(hi, v))


def _today() -> dt.date:
    return dt.datetime.now().date()


def _lookback_start(days: int) -> dt.date:
    return _today() - dt.timedelta(days=days)


@dataclass
class ScanConfig:
    apify_token: str
    mistral_key: str
    days_lookback: int = field(default_factory=lambda: _env_int("DAYS_LOOKBACK", 90, 7, 365))
    max_items_per_source: int = field(
        default_factory=lambda: _env_int("MAX_ITEMS_PER_SOURCE", 4, 3, 5)
    )
    output_xlsx: str = field(default_factory=lambda: os.getenv("OUTPUT_XLSX", "hk_beauty_device_trial_scan.xlsx"))
    mistral_model: str = field(default_factory=lambda: os.getenv("MISTRAL_MODEL", "mistral-small-latest"))
    mistral_base: str = field(
        default_factory=lambda: os.getenv("MISTRAL_API_BASE_URL", "https://api.mistral.ai/v1").rstrip("/")
    )
    actors: Dict[str, str] = field(default_factory=dict)
    google_search_actor: str = field(
        default_factory=lambda: os.getenv("APIFY_ACTOR_GOOGLE_SEARCH", "apify~google-search-scraper")
    )
    max_distributor_queries: int = field(default_factory=lambda: _env_int("MAX_DISTRIBUTOR_QUERIES", 3, 1, 6))
    max_urls_per_query: int = field(default_factory=lambda: _env_int("MAX_URLS_PER_DISTRIBUTOR_QUERY", 4, 1, 10))
    sleep_between_llm: float = field(default_factory=lambda: float(os.getenv("LLM_SLEEP_SECONDS", "2.0")))

    @staticmethod
    def load() -> "ScanConfig":
        token = os.getenv("APIFY_API_TOKEN", "").strip()
        mistral = os.getenv("MISTRAL_API_KEY", "").strip()
        actors = {
            "xhs": os.getenv("APIFY_ACTOR_XHS", "").strip(),
            "instagram": os.getenv("APIFY_ACTOR_INSTAGRAM", "").strip(),
            "tiktok": os.getenv("APIFY_ACTOR_TIKTOK", "").strip(),
            "x": os.getenv("APIFY_ACTOR_X", "").strip(),
        }
        return ScanConfig(apify_token=token, mistral_key=mistral, actors=actors)


def _session() -> requests.Session:
    s = requests.Session()
    s.trust_env = not IGNORE_SYSTEM_PROXY
    return s


def _parse_json_env(name: str) -> Dict[str, Any]:
    raw = os.getenv(name, "").strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        logging.warning("Invalid JSON in %s; ignoring.", name)
        return {}


# ---------------------------------------------------------------------------
# Apify REST helpers
# ---------------------------------------------------------------------------


def _actor_path(actor_id: str) -> str:
    """Accept `user~name` or `user/name` from copy-paste."""
    aid = actor_id.strip().replace("/", "~")
    return quote(aid, safe="")


def apify_start_run(token: str, actor_id: str, run_input: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    url = f"https://api.apify.com/v2/acts/{_actor_path(actor_id)}/runs"
    params = {"token": token, "waitForFinish": "300"}
    try:
        r = _session().post(url, params=params, json=run_input, timeout=REQUEST_TIMEOUT + 320)
        r.raise_for_status()
        return r.json()
    except requests.RequestException as exc:
        logging.error("Apify run failed for actor %s: %s", actor_id, exc)
        return None


def apify_poll_run(token: str, run_id: str, max_wait_s: int = 600) -> Tuple[str, str]:
    url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    deadline = time.time() + max_wait_s
    status = "UNKNOWN"
    dataset_id = ""
    while time.time() < deadline:
        try:
            r = _session().get(url, params={"token": token}, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            data = r.json().get("data", {})
            status = str(data.get("status", ""))
            dataset_id = str(data.get("defaultDatasetId", "") or "")
            if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
                return status, dataset_id
        except requests.RequestException as exc:
            logging.warning("Apify poll error: %s", exc)
        time.sleep(8)
    return status or "TIMED-OUT", dataset_id


def apify_fetch_dataset(token: str, dataset_id: str, limit: int) -> List[Dict[str, Any]]:
    if not dataset_id:
        return []
    url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    params = {"token": token, "clean": "true", "format": "json", "limit": str(limit)}
    try:
        r = _session().get(url, params=params, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        out = r.json()
        return out if isinstance(out, list) else []
    except requests.RequestException as exc:
        logging.error("Apify dataset fetch failed: %s", exc)
        return []


def apify_run_actor_sync(
    token: str, actor_id: str, run_input: Dict[str, Any], limit_items: int
) -> List[Dict[str, Any]]:
    if not actor_id:
        return []
    started = apify_start_run(token, actor_id, run_input)
    if not started:
        return []
    run = started.get("data", started)
    run_id = str(run.get("id", ""))
    status = str(run.get("status", ""))
    dataset_id = str(run.get("defaultDatasetId", "") or "")
    if status not in {"SUCCEEDED", "READY"} and run_id:
        status, dataset_id = apify_poll_run(token, run_id)
    if status != "SUCCEEDED":
        logging.warning("Apify actor %s finished with status=%s", actor_id, status)
        return []
    return apify_fetch_dataset(token, dataset_id, limit_items)


# ---------------------------------------------------------------------------
# Default actor inputs (best-effort; override with APIFY_INPUT_OVERRIDE_JSON)
# ---------------------------------------------------------------------------


def _default_actor_input(platform: str, terms: List[str], max_items: int) -> Dict[str, Any]:
    """
    Conservative defaults that work with several popular Store actors.
    Always validate against your actor's README and override via env JSON.
    """
    t = terms[:6]
    mi = max_items
    if platform == "tiktok":
        return {
            "searchQueries": t,
            "resultsPerPage": mi,
            "shouldDownloadVideos": False,
            "shouldDownloadCovers": False,
        }
    if platform == "instagram":
        return {
            "search": t[0] if t else "beauty device",
            "searchType": "hashtag",
            "searchLimit": 1,
            "resultsType": "posts",
            "resultsLimit": mi,
        }
    if platform == "xhs":
        return {
            "searchKeywords": t,
            "searchType": "keyword",
            "maxItems": mi,
        }
    if platform == "x":
        return {
            "searchTerms": t,
            "maxTweets": mi,
            "maxItems": mi,
        }
    return {"queries": t, "maxItems": mi}


def merge_actor_input(platform: str, base: Dict[str, Any], overrides: Dict[str, Any]) -> Dict[str, Any]:
    plat_over = overrides.get(platform) if isinstance(overrides, dict) else None
    if isinstance(plat_over, dict):
        merged = {**base, **plat_over}
        return merged
    return base


# ---------------------------------------------------------------------------
# Normalise messy Apify items → one record per post/video/tweet
# ---------------------------------------------------------------------------


@dataclass
class SocialRecord:
    platform: str
    title: str
    text: str
    url: str
    posted_at_hint: str
    engagement: str
    image_url: str
    raw: Dict[str, Any]


def _first_str(item: Dict[str, Any], keys: Tuple[str, ...]) -> str:
    for k in keys:
        v = item.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _first_http(item: Dict[str, Any], keys: Tuple[str, ...]) -> str:
    s = _first_str(item, keys)
    if s.startswith("http"):
        return s
    return ""


def flatten_text(item: Dict[str, Any]) -> str:
    parts: List[str] = []
    for k in (
        "title",
        "desc",
        "description",
        "text",
        "caption",
        "content",
        "noteTitle",
        "noteContent",
        "videoDescription",
        "fullText",
    ):
        v = item.get(k)
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    return " ".join(parts)[:6000]


def flatten_posted_at(item: Dict[str, Any]) -> str:
    return _first_str(
        item,
        (
            "createTime",
            "createdAt",
            "date",
            "timestamp",
            "publishedAt",
            "takenAt",
            "time",
            "create_time",
        ),
    )


def flatten_url(item: Dict[str, Any]) -> str:
    return _first_http(
        item,
        (
            "url",
            "postUrl",
            "webVideoUrl",
            "videoUrl",
            "link",
            "shortCodeUrl",
            "noteUrl",
            "webUrl",
        ),
    )


def flatten_image(item: Dict[str, Any]) -> str:
    u = _first_http(item, ("displayUrl", "thumbnailUrl", "coverUrl", "imageUrl", "thumbnail", "picUrl"))
    if u:
        return u
    imgs = item.get("images") or item.get("imageUrls")
    if isinstance(imgs, list):
        for el in imgs:
            if isinstance(el, str) and el.startswith("http"):
                return el
            if isinstance(el, dict):
                u2 = el.get("url") or el.get("src")
                if isinstance(u2, str) and u2.startswith("http"):
                    return u2
    return ""


def engagement_summary(item: Dict[str, Any]) -> str:
    pairs: List[str] = []
    mapping = {
        "views": ("viewCount", "playCount", "videoViewCount", "readCount"),
        "likes": ("likes", "likeCount", "diggCount", "likedCount"),
        "comments": ("comments", "commentCount"),
        "shares": ("shares", "shareCount"),
        "saves": ("collectCount", "saveCount", "bookmarkCount"),
    }
    for label, keys in mapping.items():
        for k in keys:
            v = item.get(k)
            if v not in (None, "", 0, "0"):
                pairs.append(f"{label}: {v}")
                break
    au = _first_str(item, ("author", "authorName", "nickname", "ownerUsername", "userName", "username"))
    if au:
        pairs.append(f"creator: {au}")
    return "; ".join(pairs) if pairs else "N/A"


def apify_item_to_record(platform: str, item: Dict[str, Any]) -> SocialRecord:
    text = flatten_text(item)
    title = _first_str(item, ("title", "noteTitle")) or (text[:140] + "…" if len(text) > 140 else text)
    url = flatten_url(item) or f"apify:{platform}:no-public-url"
    return SocialRecord(
        platform=platform,
        title=title,
        text=text,
        url=url,
        posted_at_hint=flatten_posted_at(item),
        engagement=engagement_summary(item),
        image_url=flatten_image(item),
        raw=item,
    )


# ---------------------------------------------------------------------------
# Mistral JSON calls
# ---------------------------------------------------------------------------


def _clean_json_blob(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\{.*\}", text, flags=re.DOTALL)
    return m.group(0).strip() if m else text


def mistral_chat_json(cfg: ScanConfig, system: str, user: str) -> Optional[Dict[str, Any]]:
    if not cfg.mistral_key:
        logging.error("MISTRAL_API_KEY is not set.")
        return None
    url = f"{cfg.mistral_base}/chat/completions"
    headers = {"Authorization": f"Bearer {cfg.mistral_key}", "Content-Type": "application/json"}
    payload = {
        "model": cfg.mistral_model,
        "temperature": 0.15,
        "max_tokens": 1800,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
    }
    try:
        r = _session().post(url, headers=headers, json=payload, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        content = data["choices"][0]["message"]["content"]
        parsed = json.loads(_clean_json_blob(content))
        time.sleep(cfg.sleep_between_llm)
        return parsed if isinstance(parsed, dict) else None
    except (requests.RequestException, KeyError, IndexError, json.JSONDecodeError, TypeError) as exc:
        logging.warning("Mistral call failed: %s", exc)
        return None


def build_extraction_prompt(cfg: ScanConfig, rec: SocialRecord) -> str:
    start = _lookback_start(cfg.days_lookback).isoformat()
    end = _today().isoformat()
    return f"""
You help a Hong Kong buyer discover **importable B2C home-use electronic beauty devices**
trending in foreign consumer social feeds (Japan/Korea/US/EU/China creators, etc.).

Today (runtime): {end}
Recency window for *first meaningful public signal*: {start} to {end} inclusive.

Social platform (Apify source): {rec.platform}
Post URL (may be missing): {rec.url}
Title: {rec.title}
Platform timestamp/metadata hint: {rec.posted_at_hint!r}
Engagement summary: {rec.engagement}

Post text (truncated):
{rec.text[:5000]}

Rules:
- Return JSON only. No markdown.
- **Reject** pure skincare/cosmetics, ambiguous gadget spam, salon-only/clinic-only/prescription
  medical devices, crowdfunding-only with no retail path, or content with no identifiable SKU-level product.
- **Accept** only if you can name **brand + product line/model** (or very clear unique descriptor).
- **Dates are critical**: compare `posted_at_hint` with textual clues ("March 2025", "下周上市", "刚入手").
  Choose `ai_best_date_iso` as your best estimate of when the *signal* (launch/review/unboxing/trend post)
  applies, within the window when possible. If unknown, set `ai_best_date_iso` to null and explain.
- `date_confidence` must be one of: high, medium, low.
- Propose up to 3 **short** Google queries to locate likely **distributors / wholesalers / regional dealers**
  (not consumer reviews). Use English plus brand native language if useful.

Return JSON with keys:
{{
  "is_candidate": true,
  "reject_reason": "",
  "product_name": "",
  "brand": "",
  "category": "",
  "functions_or_hooks": "",
  "origin_market_guess": "",
  "posted_metadata_raw": "",
  "ai_best_date_iso": "YYYY-MM-DD",
  "date_confidence": "high|medium|low",
  "date_reasoning": "",
  "hk_relevance": "High|Medium|Low",
  "hk_relevance_reason": "",
  "distributor_search_queries": ["...", "..."],
  "notes": ""
}}
""".strip()


def is_in_lookback(cfg: ScanConfig, iso_date: Optional[str]) -> bool:
    if not iso_date:
        return False
    try:
        parsed = dt.date.fromisoformat(iso_date[:10])
    except ValueError:
        return False
    return _lookback_start(cfg.days_lookback) <= parsed <= _today()


def postprocess_extraction(raw: Dict[str, Any], rec: SocialRecord) -> Dict[str, Any]:
    out = dict(raw)
    out.setdefault("is_candidate", False)
    out.setdefault("product_name", "")
    out.setdefault("brand", "")
    out.setdefault("ai_best_date_iso", None)
    out.setdefault("date_confidence", "low")
    out.setdefault("distributor_search_queries", [])
    out.setdefault("posted_metadata_raw", "")
    if not str(out.get("posted_metadata_raw", "")).strip():
        out["posted_metadata_raw"] = rec.posted_at_hint or "N/A"
    out["source_url"] = rec.url
    out["platform"] = rec.platform
    out["engagement"] = rec.engagement
    out["image_url"] = rec.image_url
    return out


# ---------------------------------------------------------------------------
# Distributor discovery via Apify Google Search Scraper
# ---------------------------------------------------------------------------


def build_google_scraper_input(queries: List[str], results_per_page: int) -> Dict[str, Any]:
    # Matches official actor example: newline-separated queries string.
    qstr = "\n".join(q for q in queries if q.strip())
    return {
        "queries": qstr,
        "maxPagesPerQuery": 1,
        "resultsPerPage": min(100, max(10, results_per_page)),
        "languageCode": "en",
        "countryCode": "us",
        "mobileResults": False,
        "includeUnfilteredResults": False,
    }


def extract_urls_from_google_dataset(items: List[Dict[str, Any]], per_query_cap: int) -> List[str]:
    urls: List[str] = []
    for block in items:
        organic = block.get("organicResults")
        if isinstance(organic, list):
            for row in organic[:per_query_cap]:
                u = row.get("url")
                if isinstance(u, str) and u.startswith("http"):
                    urls.append(u)
        # Some actors emit flat rows
        u2 = block.get("url")
        if isinstance(u2, str) and u2.startswith("http") and u2 not in urls:
            urls.append(u2)
    # de-dupe preserving order
    seen = set()
    out: List[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def find_distributor_urls(cfg: ScanConfig, queries: List[str]) -> str:
    if not cfg.apify_token or not queries:
        return ""
    actor = cfg.google_search_actor
    inp = build_google_scraper_input(queries, max(10, cfg.max_urls_per_query * 10))
    items = apify_run_actor_sync(cfg.apify_token, actor, inp, limit_items=20)
    urls = extract_urls_from_google_dataset(items, cfg.max_urls_per_query)
    return "\n".join(urls[: cfg.max_urls_per_query * cfg.max_distributor_queries])


# ---------------------------------------------------------------------------
# Excel export (with thumbnails when possible)
# ---------------------------------------------------------------------------

EXCEL_COLUMNS = [
    "Platform",
    "Brand",
    "Product",
    "Category",
    "Functions / hooks",
    "Origin guess",
    "Posted metadata (site)",
    "AI best date (ISO)",
    "Date confidence",
    "Date reasoning",
    "HK relevance",
    "HK relevance reason",
    "Engagement",
    "Source URL",
    "Distributor / B2B URLs (Apify Google)",
    "Notes",
]


def _thumb_from_url(url: str, out_dir: Path) -> Optional[Path]:
    if not url.startswith("http"):
        return None
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
        )
    }
    digest = hashlib.sha256(url.encode("utf-8", errors="ignore")).hexdigest()[:16]
    raw_path = out_dir / f"img_{digest}.raw"
    png_path = out_dir / f"img_{digest}.png"
    try:
        r = _session().get(url, headers=headers, stream=True, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        with raw_path.open("wb") as fh:
            for chunk in r.iter_content(8192):
                if chunk:
                    fh.write(chunk)
        with Image.open(raw_path) as im:
            im = ImageOps.exif_transpose(im)
            im.thumbnail((120, 120), Image.Resampling.LANCZOS)
            canvas = Image.new("RGB", (120, 120), "white")
            x = (120 - im.width) // 2
            y = (120 - im.height) // 2
            if im.mode in ("RGBA", "LA"):
                canvas.paste(im, (x, y), im.convert("RGBA"))
            else:
                canvas.paste(im.convert("RGB"), (x, y))
            canvas.save(png_path, format="PNG")
        raw_path.unlink(missing_ok=True)
        return png_path
    except (requests.RequestException, OSError, UnidentifiedImageError) as exc:
        logging.info("Thumbnail skipped (%s): %s", url, exc)
        try:
            raw_path.unlink(missing_ok=True)
        except OSError:
            pass
        return None


def write_workbook(rows: List[Dict[str, Any]], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial scan"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    thumb_col = len(EXCEL_COLUMNS) + 1
    for col, name in enumerate(EXCEL_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    th = ws.cell(row=1, column=thumb_col, value="Thumbnail")
    th.fill = hdr_fill
    th.font = hdr_font
    th.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [14, 18, 32, 18, 40, 16, 28, 16, 14, 44, 12, 36, 30, 52, 52, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[get_column_letter(thumb_col)].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}1"

    with tempfile.TemporaryDirectory(prefix="hk_beauty_scan_") as td:
        tdir = Path(td)
        for ridx, row in enumerate(rows, start=2):
            ws.row_dimensions[ridx].height = 96
            for cidx, key in enumerate(EXCEL_COLUMNS, start=1):
                ws.cell(row=ridx, column=cidx, value=row.get(key, "")).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            img_u = row.get("_image_url", "")
            if img_u:
                p = _thumb_from_url(str(img_u), tdir)
                if p:
                    xl_img = ExcelImage(str(p))
                    xl_img.width = 120
                    xl_img.height = 120
                    anchor = f"{get_column_letter(thumb_col)}{ridx}"
                    ws.add_image(xl_img, anchor)

    wb.save(path)
    logging.info("Wrote workbook: %s", path)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def scan_platform(
    cfg: ScanConfig,
    platform_key: str,
    display_name: str,
    terms: List[str],
    overrides: Dict[str, Any],
) -> List[Dict[str, Any]]:
    actor = cfg.actors.get(platform_key, "")
    if not actor:
        logging.warning("No Apify actor configured for %s — set APIFY_ACTOR_%s", display_name, platform_key.upper())
        return []

    base_in = _default_actor_input(platform_key, terms, cfg.max_items_per_source)
    run_input = merge_actor_input(platform_key, base_in, overrides)
    items = apify_run_actor_sync(cfg.apify_token, actor, run_input, limit_items=cfg.max_items_per_source * 2)
    if not items:
        logging.warning("Apify returned no items for %s", display_name)
        return []

    records = [apify_item_to_record(display_name, it) for it in items[: cfg.max_items_per_source * 3]]
    # trim to max_items after light filter: must have text
    records = [r for r in records if r.text.strip()][: cfg.max_items_per_source]

    rows: List[Dict[str, Any]] = []
    for rec in records:
        user = build_extraction_prompt(cfg, rec)
        parsed = mistral_chat_json(
            cfg,
            system="You return compact factual JSON only. Never invent prices or certifications.",
            user=user,
        )
        if not parsed:
            continue
        row = postprocess_extraction(parsed, rec)
        if not row.get("is_candidate"):
            logging.info("Rejected by LLM: %s — %s", rec.url, row.get("reject_reason", ""))
            continue
        iso = row.get("ai_best_date_iso")
        if isinstance(iso, str):
            iso = iso[:10]
        if not is_in_lookback(cfg, iso if isinstance(iso, str) else None):
            logging.info("Outside lookback or missing date: %s (%s)", row.get("product_name"), iso)
            continue

        queries = row.get("distributor_search_queries") or []
        if isinstance(queries, str):
            queries = [queries]
        queries = [str(q).strip() for q in queries if str(q).strip()][: cfg.max_distributor_queries]
        dist_urls = find_distributor_urls(cfg, queries) if queries else ""

        rows.append(
            {
                "Platform": display_name,
                "Brand": row.get("brand", ""),
                "Product": row.get("product_name", ""),
                "Category": row.get("category", ""),
                "Functions / hooks": row.get("functions_or_hooks", ""),
                "Origin guess": row.get("origin_market_guess", ""),
                "Posted metadata (site)": row.get("posted_metadata_raw", rec.posted_at_hint),
                "AI best date (ISO)": iso or "",
                "Date confidence": row.get("date_confidence", ""),
                "Date reasoning": row.get("date_reasoning", ""),
                "HK relevance": row.get("hk_relevance", ""),
                "HK relevance reason": row.get("hk_relevance_reason", ""),
                "Engagement": row.get("engagement", ""),
                "Source URL": row.get("source_url", ""),
                "Distributor / B2B URLs (Apify Google)": dist_urls,
                "Notes": row.get("notes", ""),
                "_image_url": row.get("image_url", ""),
            }
        )
        if len(rows) >= cfg.max_items_per_source:
            break

    return rows


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
    cfg = ScanConfig.load()
    overrides = _parse_json_env("APIFY_INPUT_OVERRIDE_JSON")

    if not cfg.apify_token:
        logging.error("APIFY_API_TOKEN is required.")
        return
    if not cfg.mistral_key:
        logging.error("MISTRAL_API_KEY is required for structured extraction and date reconciliation.")
        return

    terms = DEFAULT_SEARCH_TERMS
    env_terms = os.getenv("SCAN_SEARCH_TERMS_JSON", "").strip()
    if env_terms:
        try:
            loaded = json.loads(env_terms)
            if isinstance(loaded, list) and all(isinstance(x, str) for x in loaded):
                terms = loaded
        except json.JSONDecodeError:
            logging.warning("SCAN_SEARCH_TERMS_JSON invalid; using defaults.")

    all_rows: List[Dict[str, Any]] = []
    for key, label in (
        ("xhs", "Xiaohongshu (XHS)"),
        ("instagram", "Instagram"),
        ("tiktok", "TikTok"),
        ("x", "X (Twitter)"),
    ):
        logging.info("=== Scanning %s ===", label)
        try:
            all_rows.extend(scan_platform(cfg, key, label, terms, overrides))
        except Exception:
            logging.exception("Platform scan crashed: %s", label)

    if not all_rows:
        logging.warning(
            "No qualifying rows. Check actor ids, input overrides, Apify run logs, and LLM output. "
            "Creating an empty workbook shell anyway."
        )

    write_workbook(all_rows, cfg.output_xlsx)


if __name__ == "__main__":
    main()
