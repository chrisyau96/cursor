"""Foreign-market beauty-device trend scanner for Hong Kong product sourcing.

This script replaces an older monolithic scanner. It focuses specifically on
B2C home-use electronic beauty devices launched in the last few months in
foreign markets (Japan, Korea, China, Taiwan, US, UK, etc.) so that a Hong
Kong sourcing buyer can shortlist 3-5 candidates per channel for trial intake.

Pipeline
--------
1.  Social discovery via Apify actors for Xiaohongshu (XHS), TikTok,
    Instagram and X (Twitter).
2.  Web discovery via Brave Search across curated retailer / press domains.
3.  YouTube discovery via the YouTube Data API.
4.  Each raw hit is normalized into a ``Candidate`` and sent to Mistral AI
    for structured product validation.
5.  When the source page does not clearly disclose a launch date, the
    script triggers a second Mistral call that cross-references additional
    Brave snippets so the launch date is AI-validated rather than guessed
    from the page alone.
6.  For each shortlisted product, the script runs a distributor-discovery
    pass (Brave + Mistral) to surface Hong Kong / Asia distributor or
    official reseller leads via the same Brave/Mistral plumbing.
7.  Results are exported to a single Excel workbook with embedded product
    thumbnails.

All API keys can be provided via environment variables; the defaults below
match the values the user supplied for trial-run convenience.  Do NOT ship
real keys to production - move them to a secrets manager before deploying.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import logging
import os
import re
import sys
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from urllib.parse import parse_qsl, quote, urlencode, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup
from dateutil import parser as date_parser
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, UnidentifiedImageError


# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

DEFAULT_BRAVE_API_KEY = os.getenv("BRAVE_API_KEY", "BSAkcO77ShLyojFQMX8ZJ8x0aPZrUIi")
DEFAULT_MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY", "IDNIEe6IhIIIIsQPHgM8Uao6N1qwdZEr")
DEFAULT_YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY", "AIzaSyCe9-5R0ZiGYg6ZO8jhimb8bmUMqwEZ2lo")
DEFAULT_APIFY_API_TOKEN = os.getenv(
    "APIFY_API_TOKEN", "apify_api_2LIxymIcu8oSy4CrF8iXiAP5iJJmrX09ju4e"
)

# Pre-wired Apify actor IDs.  Override via env if you prefer different ones.
# These actor IDs follow the "<username>/<actor>" convention used by Apify.
APIFY_ACTORS: Dict[str, str] = {
    "Xiaohongshu": os.getenv("APIFY_ACTOR_XHS", "easyapi/xiaohongshu-search-scraper"),
    "TikTok": os.getenv("APIFY_ACTOR_TIKTOK", "clockworks/tiktok-scraper"),
    "Instagram": os.getenv("APIFY_ACTOR_INSTAGRAM", "apify/instagram-hashtag-scraper"),
    "X": os.getenv("APIFY_ACTOR_X", "apidojo/tweet-scraper"),
}

MISTRAL_API_BASE_URL = "https://api.mistral.ai/v1"
MISTRAL_MODEL = os.getenv("MISTRAL_MODEL", "mistral-small-latest")
MISTRAL_MODEL_FALLBACKS = [
    "mistral-small-latest",
    "mistral-medium-latest",
    "mistral-large-latest",
]

# Default scan settings - tweakable via CLI flags below.
DEFAULT_DAYS_LOOKBACK = 120
DEFAULT_PER_SOURCE = 3  # user wants 3-5 per source for a trial run
DEFAULT_OUTPUT_FILE = "foreign_beauty_device_scan.xlsx"

REQUEST_TIMEOUT_SECONDS = 15
PAGE_TEXT_LIMIT = 9000
SLEEP_BETWEEN_API_CALLS_SECONDS = 1.5
MISTRAL_RATE_LIMIT_WAIT_SECONDS = 65
MIN_SOCIAL_ENGAGEMENT_THRESHOLD = 80
APIFY_RUN_POLL_INTERVAL_SECONDS = 8
APIFY_RUN_MAX_WAIT_SECONDS = 240
IGNORE_SYSTEM_PROXY = True

# Curated retailer / press domains that consistently surface foreign
# beauty-device launches.  Each entry produces a couple of Brave queries.
WEB_SOURCES: List[Dict[str, Any]] = [
    {
        "name": "CurrentBody",
        "domain": "www.currentbody.com",
        "market": "International / UK / US",
        "keywords": ["beauty device new launch", "LED mask new", "RF device new"],
    },
    {
        "name": "Sephora US",
        "domain": "www.sephora.com",
        "market": "US",
        "keywords": ["beauty device new", "skincare device new launch"],
    },
    {
        "name": "Amazon Japan",
        "domain": "www.amazon.co.jp",
        "market": "Japan",
        "keywords": ["美顔器 新商品", "美容家電 新商品", "脱毛器 新商品"],
    },
    {
        "name": "Rakuten Japan",
        "domain": "www.rakuten.co.jp",
        "market": "Japan",
        "keywords": ["美顔器 新商品", "LED マスク 新商品"],
    },
    {
        "name": "@cosme Japan",
        "domain": "www.cosme.net",
        "market": "Japan",
        "keywords": ["美顔器 ランキング", "美容家電 新商品"],
    },
    {
        "name": "Olive Young Global",
        "domain": "global.oliveyoung.com",
        "market": "Korea / Global",
        "keywords": ["beauty device new", "Korean beauty device"],
    },
    {
        "name": "Hwahae",
        "domain": "www.hwahae.com",
        "market": "Korea",
        "keywords": ["뷰티 디바이스 신제품", "LED 마스크 신제품"],
    },
    {
        "name": "momo Shopping Taiwan",
        "domain": "www.momoshop.com.tw",
        "market": "Taiwan",
        "keywords": ["美容儀 新品", "美容家電 新品"],
    },
    {
        "name": "BeautyMatter",
        "domain": "beautymatter.com",
        "market": "International",
        "keywords": ["beauty device launch", "beauty tech launch"],
    },
    {
        "name": "PR Newswire",
        "domain": "www.prnewswire.com",
        "market": "International / US",
        "keywords": ["beauty device launch", "LED mask launch"],
    },
]

YOUTUBE_QUERIES = [
    "beauty device new launch",
    "home beauty device review",
    "LED face mask new",
    "RF facial device review",
    "Korean beauty device review",
    "美顔器 新商品",
    "美容家電 新商品",
]

SOCIAL_SEARCH_TERMS = [
    "beauty device new launch",
    "LED face mask",
    "RF facial device",
    "microcurrent device",
    "IPL hair removal device",
    "美容儀",
    "美顔器",
    "뷰티 디바이스",
]

REQUIRED_COLUMNS = [
    "Source Type",
    "Source Name",
    "Country / Market",
    "Product Category",
    "Product Name",
    "Brand",
    "Product Image",
    "Key Functions / Selling Points",
    "Launch Date (AI Verified)",
    "Date Source / Evidence",
    "Price",
    "Supplier / Brand Information",
    "Engagement / Popularity Signal",
    "Relevance to HK Sourcing",
    "Distributor / Reseller Leads",
    "Source URL",
]


# ---------------------------------------------------------------------------
# DATACLASSES
# ---------------------------------------------------------------------------


@dataclass
class Candidate:
    source_type: str
    source_name: str
    country_or_market: str
    title: str
    url: str
    snippet: str = ""
    source_date_hint: str = ""
    page_text: str = ""
    image_url_hint: str = ""
    engagement_hint: str = ""
    raw_payload: Optional[Dict[str, Any]] = None


@dataclass
class ScannerConfig:
    brave_key: str
    mistral_key: str
    youtube_key: str
    apify_token: str
    per_source: int
    days_lookback: int
    output_file: str
    enabled_channels: List[str]
    dry_run: bool = False
    skip_distributor: bool = False


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------


def setup_logging(verbose: bool = False) -> None:
    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    root = logging.getLogger()
    root.setLevel(logging.DEBUG if verbose else logging.INFO)
    root.handlers.clear()
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(formatter)
    root.addHandler(handler)


def redact_url(url: str) -> str:
    sensitive = {"key", "token", "api_key", "apikey", "x-subscription-token"}
    try:
        parts = urlsplit(url)
        query = urlencode(
            [
                (k, "***" if k.lower() in sensitive else v)
                for k, v in parse_qsl(parts.query, keep_blank_values=True)
            ]
        )
        return urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment))
    except Exception:  # noqa: BLE001
        return re.sub(r"(?i)(key|token|api_key|apikey)=([^&\s]+)", r"\1=***", url)


def redact_text(text: str) -> str:
    text = re.sub(r"(?i)(key|token|api_key|apikey)=([^&\s]+)", r"\1=***", text)
    return re.sub(r"(?i)(bearer\s+)[a-z0-9_\-\.]+", r"\1***", text)


def today() -> dt.date:
    return dt.datetime.now().date()


def start_window(days_lookback: int) -> dt.date:
    return today() - dt.timedelta(days=days_lookback)


def http_request(
    method: str,
    url: str,
    *,
    headers: Optional[Dict[str, str]] = None,
    params: Optional[Dict[str, Any]] = None,
    json_body: Optional[Dict[str, Any]] = None,
    timeout: int = REQUEST_TIMEOUT_SECONDS,
    retries: int = 2,
    parse_json: bool = True,
) -> Optional[Any]:
    safe = redact_url(url)
    for attempt in range(retries + 1):
        try:
            with requests.Session() as session:
                session.trust_env = not IGNORE_SYSTEM_PROXY
                response = session.request(
                    method,
                    url,
                    headers=headers,
                    params=params,
                    json=json_body,
                    timeout=timeout,
                )
            if response.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                retry_after = response.headers.get("Retry-After")
                if response.status_code == 429:
                    wait = (
                        int(retry_after)
                        if retry_after and retry_after.isdigit()
                        else MISTRAL_RATE_LIMIT_WAIT_SECONDS
                    )
                else:
                    wait = 2 + attempt * 3
                logging.warning("HTTP %s from %s; retry in %ss", response.status_code, safe, wait)
                time.sleep(wait)
                continue
            response.raise_for_status()
            return response.json() if parse_json else response.text
        except requests.RequestException as exc:
            err = redact_text(str(exc))
            body = ""
            r = getattr(exc, "response", None)
            if r is not None and getattr(r, "text", ""):
                body = f" | body: {redact_text(r.text[:400])}"
            if attempt < retries:
                wait = 2 + attempt * 3
                logging.warning("Request failed for %s: %s%s; retry in %ss", safe, err, body, wait)
                time.sleep(wait)
                continue
            logging.error("Request gave up for %s: %s%s", safe, err, body)
        except json.JSONDecodeError as exc:
            logging.error("Invalid JSON from %s: %s", safe, exc)
            return None
    return None


def clean_json_text(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"```$", "", text).strip()
    match = re.search(r"[\{\[].*[\}\]]", text, flags=re.DOTALL)
    return match.group(0).strip() if match else text


def safe_filename(url: str, suffix: str) -> str:
    digest = hashlib.sha256(url.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return f"img_{digest}{suffix}"


# ---------------------------------------------------------------------------
# BRAVE SEARCH
# ---------------------------------------------------------------------------


def brave_search(api_key: str, query: str, count: int = 5) -> List[Dict[str, Any]]:
    if not api_key:
        logging.warning("Brave API key missing; skipping query: %s", query)
        return []
    url = "https://api.search.brave.com/res/v1/web/search"
    headers = {"Accept": "application/json", "X-Subscription-Token": api_key}
    params = {
        "q": query,
        "count": min(max(count, 1), 20),
        "result_filter": "web",
        "text_decorations": "false",
        "spellcheck": "true",
    }
    data = http_request("GET", url, headers=headers, params=params)
    if not isinstance(data, dict):
        return []
    return data.get("web", {}).get("results", []) or []


def fetch_page_content(url: str) -> Tuple[str, str, str]:
    """Returns (text, image_url, raw_html_date_hint)."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        )
    }
    try:
        with requests.Session() as session:
            session.trust_env = not IGNORE_SYSTEM_PROXY
            response = session.get(url, headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
    except requests.RequestException as exc:
        logging.info("Page unavailable for %s: %s", url, exc)
        return "", "", ""

    soup = BeautifulSoup(response.text, "html.parser")
    image_url = extract_image_from_soup(soup)
    date_hint = extract_date_hint_from_soup(soup)

    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    text = re.sub(r"\s+", " ", soup.get_text(separator=" ", strip=True)).strip()
    return text[:PAGE_TEXT_LIMIT], image_url, date_hint


def extract_image_from_soup(soup: BeautifulSoup) -> str:
    for tag_name, attrs in [
        ("meta", {"property": "og:image"}),
        ("meta", {"name": "og:image"}),
        ("meta", {"name": "twitter:image"}),
        ("meta", {"property": "twitter:image"}),
    ]:
        tag = soup.find(tag_name, attrs=attrs)
        if tag and tag.get("content"):
            return str(tag["content"]).strip()
    img = soup.find("img")
    if img and img.get("src"):
        return str(img["src"]).strip()
    return ""


def extract_date_hint_from_soup(soup: BeautifulSoup) -> str:
    selectors = [
        ("meta", {"property": "article:published_time"}),
        ("meta", {"name": "pubdate"}),
        ("meta", {"name": "publishdate"}),
        ("meta", {"name": "date"}),
        ("meta", {"itemprop": "datePublished"}),
    ]
    for tag, attrs in selectors:
        node = soup.find(tag, attrs=attrs)
        if node and node.get("content"):
            return str(node["content"]).strip()
    time_node = soup.find("time")
    if time_node:
        if time_node.get("datetime"):
            return str(time_node["datetime"]).strip()
        if time_node.text:
            return time_node.text.strip()
    return ""


# ---------------------------------------------------------------------------
# MISTRAL
# ---------------------------------------------------------------------------


class MistralClient:
    def __init__(self, api_key: str, model: str = MISTRAL_MODEL):
        self.api_key = api_key
        self.model = model
        self.available: Optional[bool] = None
        self.failure_reason = ""

    def _headers(self) -> Dict[str, str]:
        return {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }

    def preflight(self) -> bool:
        if self.available is not None:
            return self.available
        if not self.api_key:
            self.available = False
            self.failure_reason = "Mistral API key missing"
            return False

        url = f"{MISTRAL_API_BASE_URL.rstrip('/')}/chat/completions"
        payload = {
            "messages": [
                {"role": "system", "content": "Return JSON only."},
                {"role": "user", "content": 'Return exactly {"ok": true}'},
            ],
            "temperature": 0,
            "max_tokens": 30,
            "response_format": {"type": "json_object"},
        }
        errors = []
        models = []
        for m in [self.model, *MISTRAL_MODEL_FALLBACKS]:
            if m not in models:
                models.append(m)
        for model in models:
            try:
                with requests.Session() as session:
                    session.trust_env = not IGNORE_SYSTEM_PROXY
                    r = session.post(
                        url,
                        headers=self._headers(),
                        json={**payload, "model": model},
                        timeout=REQUEST_TIMEOUT_SECONDS,
                    )
                if r.ok:
                    self.model = model
                    self.available = True
                    logging.info("Mistral preflight ok (model=%s)", model)
                    return True
                errors.append(f"{model}: HTTP {r.status_code}")
                if r.status_code not in {400, 404}:
                    break
            except requests.RequestException as exc:
                errors.append(f"{model}: {redact_text(str(exc))}")
                break
        self.available = False
        self.failure_reason = "Mistral preflight failed: " + " | ".join(errors)
        logging.error(self.failure_reason)
        return False

    def chat_json(self, system: str, user: str, *, max_tokens: int = 1400) -> Optional[Any]:
        if not self.preflight():
            return None
        url = f"{MISTRAL_API_BASE_URL.rstrip('/')}/chat/completions"
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            "temperature": 0.1,
            "top_p": 0.85,
            "max_tokens": max_tokens,
            "response_format": {"type": "json_object"},
        }
        data = http_request(
            "POST",
            url,
            headers=self._headers(),
            json_body=payload,
            retries=1,
        )
        if not data:
            return None
        try:
            text = data["choices"][0]["message"]["content"]
            return json.loads(clean_json_text(text))
        except (KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
            logging.warning("Mistral JSON parse failed: %s", exc)
            return None


# ---------------------------------------------------------------------------
# AI PROMPTS
# ---------------------------------------------------------------------------


PRODUCT_SYSTEM_PROMPT = (
    "You are a strict JSON extraction engine helping a Hong Kong sourcing buyer "
    "shortlist newly-launched B2C home-use electronic beauty devices from foreign "
    "markets. Return exactly one JSON object and no prose."
)


def build_product_prompt(candidate: Candidate, days_lookback: int) -> str:
    current = today().isoformat()
    earliest = start_window(days_lookback).isoformat()
    return f"""
TASK
Validate whether the candidate source describes ONE specific, newly launched
or recently discussed B2C home-use electronic beauty device suitable for
trial sourcing in Hong Kong.

Current date: {current}
Allowed source-date window: {earliest} to {current} (inclusive)

INCLUSION RULES
- Specific identifiable product with brand AND product/series/model name.
- B2C home-use electronic beauty device (LED mask, RF, microcurrent, EMS, IPL,
  hair-growth helmet, scalp device, ultrasonic cleanser, etc.).
- Discussed / launched / reviewed within the date window above.

EXCLUSION RULES
- Pure cosmetics, skincare creams, supplements.
- Generic categories ("LED mask" with no brand/model).
- Clinic-only, salon-only, medical-grade or prescription devices.
- Pure concept / prototype / crowdfunding-only items.
- Do not fabricate launch date, supplier, price, HK availability, or popularity.

OUTPUT JSON (return ONLY this object):
{{
  "is_valid_product": true | false,
  "rejection_reason": "",
  "product_category": "...",
  "product_name": "...",
  "brand": "...",
  "key_functions": "...",
  "page_disclosed_date": "YYYY-MM-DD or empty",
  "page_disclosed_date_evidence": "short quote from page or empty",
  "price": "...",
  "supplier_or_brand_info": "...",
  "engagement_or_popularity_signal": "...",
  "relevance_rating": "High" | "Medium" | "Low",
  "relevance_reason": "...",
  "image_url": "..."
}}

CANDIDATE SOURCE
source_type: {candidate.source_type}
source_name: {candidate.source_name}
country_or_market: {candidate.country_or_market}
title: {candidate.title}
url: {candidate.url}
snippet: {candidate.snippet}
source_date_hint: {candidate.source_date_hint}
image_url_hint: {candidate.image_url_hint}
engagement_hint: {candidate.engagement_hint}
page_text: {candidate.page_text}
raw_payload_excerpt: {json.dumps(candidate.raw_payload or {}, ensure_ascii=False)[:2500]}
""".strip()


DATE_SYSTEM_PROMPT = (
    "You are a meticulous research assistant. You estimate the most likely "
    "real-world launch date of a beauty device by triangulating multiple web "
    "snippets. Return strict JSON only."
)


def build_date_prompt(
    brand: str,
    product_name: str,
    page_date: str,
    page_evidence: str,
    extra_snippets: List[Dict[str, str]],
    days_lookback: int,
) -> str:
    current = today().isoformat()
    earliest = start_window(days_lookback).isoformat()
    snippets_text = json.dumps(extra_snippets, ensure_ascii=False)[:4000]
    return f"""
You are estimating the launch date of the beauty device below.

Brand: {brand}
Product: {product_name}
Page-disclosed date (may be missing or unreliable): {page_date}
Page evidence quote: {page_evidence}

Additional Brave web snippets (titles + descriptions + page_age):
{snippets_text}

Current date: {current}
Acceptable launch window: {earliest} to {current}

Return JSON:
{{
  "ai_launch_date": "YYYY-MM-DD or YYYY-MM or empty",
  "confidence": "High" | "Medium" | "Low",
  "evidence": "1-2 sentences citing which snippet you relied on",
  "within_window": true | false
}}

Rules:
- Prefer the EARLIEST credible launch / release / available-from date from the
  snippets if it falls inside the acceptable window.
- If the snippets clearly indicate a launch outside the window, set
  within_window=false and still return your best date estimate.
- If you genuinely cannot tell, leave ai_launch_date empty and confidence Low.
- Do not invent a date.
""".strip()


DISTRIBUTOR_SYSTEM_PROMPT = (
    "You extract distributor / reseller / official-store leads for a Hong Kong "
    "sourcing buyer. Return strict JSON only."
)


def build_distributor_prompt(
    brand: str,
    product_name: str,
    snippets: List[Dict[str, str]],
) -> str:
    snippets_text = json.dumps(snippets, ensure_ascii=False)[:4500]
    return f"""
You are identifying distributor, reseller, official-store, or B2B sourcing
contacts for the following product. Prefer Hong Kong / Greater China leads,
then Asia, then global.

Brand: {brand}
Product: {product_name}

Brave snippets:
{snippets_text}

Return JSON:
{{
  "leads": [
    {{
      "name": "company or store name",
      "type": "Official Brand Site | Distributor | Reseller | Marketplace | Unknown",
      "region": "HK | China | Asia | Global | ...",
      "url": "https://..."
    }}
  ],
  "summary": "1-2 sentence buyer-facing summary"
}}

Rules:
- Up to 5 leads, most relevant first.
- Skip clearly irrelevant pages (review blogs, news articles, forums).
- Never fabricate URLs - only use URLs that appear in the snippets.
- If no useful leads, return leads: [] and summary explaining why.
""".strip()


# ---------------------------------------------------------------------------
# PRODUCT VALIDATION + DATE TRIANGULATION + DISTRIBUTOR LOOKUP
# ---------------------------------------------------------------------------


def parse_date(text: str) -> Optional[dt.date]:
    if not text:
        return None
    text = text.strip()
    if not text or text.lower() in {"n/a", "none", "null", "empty"}:
        return None
    cleaned = (
        text.replace("年", "-")
        .replace("月", "-")
        .replace("日", "")
        .replace("/", "-")
        .replace(".", "-")
    )
    # YYYY-MM only → coerce to 1st of month
    if re.fullmatch(r"\d{4}-\d{1,2}", cleaned):
        cleaned = cleaned + "-01"
    try:
        return date_parser.parse(cleaned, fuzzy=True).date()
    except (ValueError, OverflowError, TypeError):
        return None


def validate_candidate(
    mistral: MistralClient,
    candidate: Candidate,
    days_lookback: int,
) -> Optional[Dict[str, Any]]:
    prompt = build_product_prompt(candidate, days_lookback)
    data = mistral.chat_json(PRODUCT_SYSTEM_PROMPT, prompt)
    time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS)
    if not data or not isinstance(data, dict):
        return None
    if not data.get("is_valid_product"):
        logging.info(
            "Mistral rejected %s: %s",
            candidate.url,
            (data.get("rejection_reason") or "")[:160],
        )
        return None
    return data


def triangulate_launch_date(
    mistral: MistralClient,
    brave_key: str,
    extracted: Dict[str, Any],
    page_date: str,
    page_evidence: str,
    days_lookback: int,
) -> Dict[str, Any]:
    brand = (extracted.get("brand") or "").strip()
    product_name = (extracted.get("product_name") or "").strip()
    if not product_name:
        return {"ai_launch_date": page_date, "confidence": "Low", "evidence": "", "within_window": False}

    query = f'"{brand}" "{product_name}" launch OR release OR review'.strip()
    snippets_raw = brave_search(brave_key, query, count=6)
    snippets = []
    for r in snippets_raw:
        snippets.append(
            {
                "title": r.get("title", "")[:200],
                "description": r.get("description", "")[:400],
                "page_age": r.get("page_age", "") or r.get("age", ""),
                "url": r.get("url", ""),
            }
        )

    if not snippets and page_date:
        parsed = parse_date(page_date)
        within = bool(parsed and start_window(days_lookback) <= parsed <= today())
        return {
            "ai_launch_date": page_date,
            "confidence": "Low",
            "evidence": page_evidence,
            "within_window": within,
        }

    prompt = build_date_prompt(brand, product_name, page_date, page_evidence, snippets, days_lookback)
    data = mistral.chat_json(DATE_SYSTEM_PROMPT, prompt, max_tokens=400)
    time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS)
    if not data or not isinstance(data, dict):
        return {
            "ai_launch_date": page_date,
            "confidence": "Low",
            "evidence": page_evidence,
            "within_window": bool(parse_date(page_date)),
        }
    return data


def lookup_distributors(
    mistral: MistralClient,
    brave_key: str,
    extracted: Dict[str, Any],
) -> Dict[str, Any]:
    brand = (extracted.get("brand") or "").strip()
    product_name = (extracted.get("product_name") or "").strip()
    if not product_name and not brand:
        return {"leads": [], "summary": "No brand/product to query"}

    queries = [
        f'"{brand}" "{product_name}" distributor Hong Kong',
        f'"{brand}" "{product_name}" official store Asia',
        f'"{brand}" distributor OR reseller Hong Kong OR China',
    ]
    snippets: List[Dict[str, str]] = []
    seen_urls = set()
    for q in queries:
        for r in brave_search(brave_key, q, count=4):
            url = r.get("url", "")
            if url in seen_urls or not url:
                continue
            seen_urls.add(url)
            snippets.append(
                {
                    "title": r.get("title", "")[:200],
                    "description": r.get("description", "")[:400],
                    "url": url,
                }
            )
        time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS / 2)

    if not snippets:
        return {"leads": [], "summary": "No Brave results"}

    prompt = build_distributor_prompt(brand, product_name, snippets)
    data = mistral.chat_json(DISTRIBUTOR_SYSTEM_PROMPT, prompt, max_tokens=700)
    time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS)
    if not data or not isinstance(data, dict):
        return {"leads": [], "summary": "Mistral did not return distributor JSON"}
    leads = data.get("leads") or []
    return {"leads": leads[:5], "summary": data.get("summary", "")}


# ---------------------------------------------------------------------------
# CANDIDATE BUILDERS
# ---------------------------------------------------------------------------


def build_web_candidates(brave_key: str, source: Dict[str, Any], per_source: int) -> List[Candidate]:
    candidates: List[Candidate] = []
    seen = set()
    for keyword in source["keywords"][:2]:
        query = f"site:{source['domain']} {keyword}"
        for r in brave_search(brave_key, query, count=per_source * 2):
            url = r.get("url") or r.get("link") or ""
            if not url or url in seen:
                continue
            seen.add(url)
            candidates.append(
                Candidate(
                    source_type="Web",
                    source_name=source["name"],
                    country_or_market=source["market"],
                    title=r.get("title", ""),
                    url=url,
                    snippet=r.get("description", ""),
                    source_date_hint=r.get("age", "") or r.get("page_age", ""),
                    raw_payload=r,
                )
            )
            if len(candidates) >= per_source * 2:
                break
        time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS)
    return candidates


def youtube_candidates(api_key: str, days_lookback: int, per_query: int = 4) -> List[Candidate]:
    if not api_key:
        logging.warning("YouTube API key missing; skipping YouTube channel")
        return []
    published_after = (
        dt.datetime.combine(start_window(days_lookback), dt.time.min).isoformat("T") + "Z"
    )

    video_ids: Dict[str, Dict[str, Any]] = {}
    for query in YOUTUBE_QUERIES:
        params = {
            "part": "snippet",
            "q": query,
            "type": "video",
            "publishedAfter": published_after,
            "maxResults": per_query,
            "order": "relevance",
            "key": api_key,
            "safeSearch": "none",
        }
        data = http_request("GET", "https://www.googleapis.com/youtube/v3/search", params=params)
        if not isinstance(data, dict):
            continue
        for item in data.get("items", []):
            vid = item.get("id", {}).get("videoId")
            if vid:
                video_ids[vid] = item
        time.sleep(SLEEP_BETWEEN_API_CALLS_SECONDS)

    if not video_ids:
        return []

    params = {
        "part": "snippet,statistics",
        "id": ",".join(video_ids),
        "key": api_key,
        "maxResults": 50,
    }
    details = http_request("GET", "https://www.googleapis.com/youtube/v3/videos", params=params)
    if not isinstance(details, dict):
        return []

    candidates: List[Candidate] = []
    for item in details.get("items", []):
        snip = item.get("snippet", {})
        stats = item.get("statistics", {})
        vid = item.get("id", "")
        thumbs = snip.get("thumbnails", {})
        thumb_url = (
            thumbs.get("maxres", {}).get("url")
            or thumbs.get("high", {}).get("url")
            or thumbs.get("medium", {}).get("url")
            or thumbs.get("default", {}).get("url")
            or ""
        )
        engagement = (
            f"Views: {stats.get('viewCount', 'N/A')}; "
            f"Likes: {stats.get('likeCount', 'N/A')}; "
            f"Comments: {stats.get('commentCount', 'N/A')}"
        )
        candidates.append(
            Candidate(
                source_type="YouTube",
                source_name="YouTube",
                country_or_market="International",
                title=snip.get("title", ""),
                url=f"https://www.youtube.com/watch?v={vid}",
                snippet=snip.get("description", "")[:2000],
                source_date_hint=snip.get("publishedAt", ""),
                image_url_hint=thumb_url,
                engagement_hint=engagement,
                raw_payload=item,
            )
        )
    return candidates


# ---------------------------------------------------------------------------
# APIFY (XHS / TikTok / Instagram / X)
# ---------------------------------------------------------------------------


def apify_input_for_platform(platform: str, max_items: int) -> Dict[str, Any]:
    """Build a per-platform Apify actor input.

    Each platform's actor expects slightly different keys.  We include the
    union of common keys so most actor variants accept the payload.
    """
    terms = SOCIAL_SEARCH_TERMS
    if platform == "Xiaohongshu":
        return {
            "keywords": terms,
            "searchTerms": terms,
            "queries": terms,
            "maxItems": max_items,
            "resultsLimit": max_items,
            "sort": "latest",
        }
    if platform == "TikTok":
        hashtags = [t.replace(" ", "") for t in terms[:6]]
        return {
            "searchQueries": terms,
            "hashtags": hashtags,
            "resultsPerPage": max_items,
            "maxItems": max_items,
            "shouldDownloadVideos": False,
            "shouldDownloadCovers": False,
            "shouldDownloadSubtitles": False,
        }
    if platform == "Instagram":
        hashtags = [re.sub(r"[^A-Za-z0-9]", "", t) for t in terms[:6] if re.search(r"[A-Za-z]", t)]
        if not hashtags:
            hashtags = ["beautydevice", "ledmask", "skincaredevice"]
        return {
            "hashtags": hashtags,
            "search": "beauty device",
            "resultsLimit": max_items,
            "resultsType": "posts",
        }
    if platform == "X":
        return {
            "searchTerms": terms,
            "queries": terms,
            "maxItems": max_items,
            "sort": "Latest",
            "tweetLanguage": "any",
        }
    return {"searchTerms": terms, "maxItems": max_items}


def run_apify_actor(api_token: str, platform: str, actor_id: str, max_items: int) -> List[Dict[str, Any]]:
    if not api_token or not actor_id:
        logging.warning("Apify token or actor id missing for %s; skipping", platform)
        return []

    encoded_actor = quote(actor_id, safe="")
    actor_input = apify_input_for_platform(platform, max_items)
    run_url = f"https://api.apify.com/v2/acts/{encoded_actor}/runs"
    params = {"token": api_token, "waitForFinish": 60}
    run_data = http_request(
        "POST",
        run_url,
        params=params,
        json_body=actor_input,
        timeout=120,
        retries=1,
    )
    if not isinstance(run_data, dict):
        return []
    run = run_data.get("data", run_data)
    status = run.get("status", "")
    run_id = run.get("id", "")
    dataset_id = run.get("defaultDatasetId", "")

    if status not in {"SUCCEEDED", "READY"} and run_id:
        status_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
        elapsed = 0
        while elapsed < APIFY_RUN_MAX_WAIT_SECONDS:
            time.sleep(APIFY_RUN_POLL_INTERVAL_SECONDS)
            elapsed += APIFY_RUN_POLL_INTERVAL_SECONDS
            poll = http_request("GET", status_url, params={"token": api_token}, retries=0)
            run = poll.get("data", {}) if isinstance(poll, dict) else {}
            status = run.get("status", "")
            dataset_id = run.get("defaultDatasetId", dataset_id)
            if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
                break

    if status != "SUCCEEDED":
        logging.warning("Apify actor for %s finished with status %s", platform, status or "unknown")
        return []
    if not dataset_id:
        logging.warning("Apify actor for %s has no dataset id", platform)
        return []

    items_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    items = http_request(
        "GET",
        items_url,
        params={"token": api_token, "clean": "true", "format": "json", "limit": max_items * 2},
    )
    return items if isinstance(items, list) else []


def social_market_for(platform: str) -> str:
    return {
        "Xiaohongshu": "Mainland China / Chinese-speaking",
        "TikTok": "International / US / SEA",
        "Instagram": "International",
        "X": "International",
    }.get(platform, "International")


def social_text(item: Dict[str, Any]) -> str:
    parts = []
    for k in ["title", "caption", "description", "text", "desc", "content", "videoDescription", "fullText"]:
        v = item.get(k)
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    return " ".join(parts)[:4000]


def social_url(item: Dict[str, Any]) -> str:
    for k in ["url", "postUrl", "webVideoUrl", "videoUrl", "link", "shortCodeUrl", "tweetUrl"]:
        v = item.get(k)
        if isinstance(v, str) and v.startswith("http"):
            return v
    return ""


def social_image(item: Dict[str, Any]) -> str:
    for k in ["displayUrl", "thumbnailUrl", "coverUrl", "image", "imageUrl", "thumbnail", "videoCover"]:
        v = item.get(k)
        if isinstance(v, str) and v.startswith("http"):
            return v
        if isinstance(v, list):
            for entry in v:
                if isinstance(entry, str) and entry.startswith("http"):
                    return entry
                if isinstance(entry, dict):
                    u = entry.get("url") or entry.get("src")
                    if isinstance(u, str) and u.startswith("http"):
                        return u
    return ""


def social_date(item: Dict[str, Any]) -> str:
    for k in ["date", "createdAt", "timestamp", "publishedAt", "takenAt", "createTime", "uploadDate"]:
        v = item.get(k)
        if v:
            return str(v)
    return ""


def social_engagement_count(item: Dict[str, Any]) -> int:
    total = 0
    for k in [
        "likes", "likeCount", "diggCount", "comments", "commentCount",
        "shares", "shareCount", "views", "viewCount", "playCount",
        "collectCount", "saveCount", "favouriteCount",
    ]:
        v = item.get(k)
        try:
            if v is not None:
                total += int(str(v).replace(",", ""))
        except ValueError:
            continue
    return total


def social_engagement_text(item: Dict[str, Any]) -> str:
    out = []
    for label, keys in {
        "Views": ["views", "viewCount", "playCount"],
        "Likes": ["likes", "likeCount", "diggCount", "favouriteCount"],
        "Comments": ["comments", "commentCount"],
        "Shares": ["shares", "shareCount"],
        "Saves": ["collectCount", "saveCount"],
        "Creator": ["author", "authorName", "username", "ownerUsername", "channelName"],
    }.items():
        for k in keys:
            v = item.get(k)
            if v not in (None, ""):
                out.append(f"{label}: {v}")
                break
    return "; ".join(out) if out else "N/A"


def apify_items_to_candidates(platform: str, items: List[Dict[str, Any]]) -> List[Candidate]:
    cands: List[Candidate] = []
    for item in items:
        if social_engagement_count(item) < MIN_SOCIAL_ENGAGEMENT_THRESHOLD:
            continue
        text = social_text(item)
        url = social_url(item)
        if not text and not url:
            continue
        cands.append(
            Candidate(
                source_type="Social",
                source_name=platform,
                country_or_market=social_market_for(platform),
                title=(item.get("title") or text)[:200],
                url=url or f"apify://{platform}",
                snippet=text,
                source_date_hint=social_date(item),
                image_url_hint=social_image(item),
                engagement_hint=social_engagement_text(item),
                raw_payload=item,
            )
        )
    return cands


# ---------------------------------------------------------------------------
# ROW BUILDING
# ---------------------------------------------------------------------------


def empty_row(source_type: str, source_name: str, market: str, reason: str) -> Dict[str, Any]:
    return {
        "Source Type": source_type,
        "Source Name": source_name,
        "Country / Market": market,
        "Product Category": "N/A",
        "Product Name": reason,
        "Brand": "N/A",
        "Product Image": "N/A",
        "Key Functions / Selling Points": "N/A",
        "Launch Date (AI Verified)": "N/A",
        "Date Source / Evidence": "N/A",
        "Price": "N/A",
        "Supplier / Brand Information": "N/A",
        "Engagement / Popularity Signal": "N/A",
        "Relevance to HK Sourcing": "N/A",
        "Distributor / Reseller Leads": "N/A",
        "Source URL": "N/A",
        "_image_url": "",
    }


def build_row(
    candidate: Candidate,
    extracted: Dict[str, Any],
    date_result: Dict[str, Any],
    distributor_result: Dict[str, Any],
) -> Dict[str, Any]:
    leads = distributor_result.get("leads", []) or []
    leads_text = "\n".join(
        f"- {lead.get('name', 'N/A')} ({lead.get('type', 'N/A')}, {lead.get('region', 'N/A')}): {lead.get('url', '')}"
        for lead in leads
    )
    summary = distributor_result.get("summary", "")
    if summary:
        leads_text = (leads_text + "\n\nSummary: " + summary).strip()
    if not leads_text:
        leads_text = summary or "No distributor leads found"

    relevance = "{} - {}".format(
        extracted.get("relevance_rating", "Low"),
        extracted.get("relevance_reason", "N/A"),
    )

    launch_date = date_result.get("ai_launch_date") or extracted.get("page_disclosed_date") or "N/A"
    date_evidence = (
        date_result.get("evidence")
        or extracted.get("page_disclosed_date_evidence")
        or "N/A"
    )
    confidence = date_result.get("confidence", "Low")
    date_evidence = f"[{confidence}] {date_evidence}"

    return {
        "Source Type": candidate.source_type,
        "Source Name": candidate.source_name,
        "Country / Market": candidate.country_or_market,
        "Product Category": extracted.get("product_category", "N/A"),
        "Product Name": extracted.get("product_name", "N/A"),
        "Brand": extracted.get("brand", "N/A"),
        "Product Image": "",
        "Key Functions / Selling Points": extracted.get("key_functions", "N/A"),
        "Launch Date (AI Verified)": launch_date,
        "Date Source / Evidence": date_evidence,
        "Price": extracted.get("price", "Not disclosed"),
        "Supplier / Brand Information": extracted.get("supplier_or_brand_info", "N/A"),
        "Engagement / Popularity Signal": extracted.get(
            "engagement_or_popularity_signal", candidate.engagement_hint or "N/A"
        ),
        "Relevance to HK Sourcing": relevance,
        "Distributor / Reseller Leads": leads_text,
        "Source URL": candidate.url,
        "_image_url": extracted.get("image_url") or candidate.image_url_hint,
    }


# ---------------------------------------------------------------------------
# PROCESSING LOOPS
# ---------------------------------------------------------------------------


def process_candidates(
    cfg: ScannerConfig,
    mistral: MistralClient,
    candidates: List[Candidate],
    *,
    fetch_page: bool,
) -> List[Dict[str, Any]]:
    """Validate + enrich a list of candidates. Returns row dicts."""
    rows: List[Dict[str, Any]] = []
    for cand in candidates:
        if len(rows) >= cfg.per_source:
            break
        if fetch_page and cand.url and cand.url.startswith("http"):
            text, image, page_date = fetch_page_content(cand.url)
            cand.page_text = text
            cand.image_url_hint = cand.image_url_hint or image
            if not cand.source_date_hint:
                cand.source_date_hint = page_date

        extracted = validate_candidate(mistral, cand, cfg.days_lookback)
        if not extracted:
            continue

        page_date = extracted.get("page_disclosed_date", "")
        page_evidence = extracted.get("page_disclosed_date_evidence", "")
        date_result = triangulate_launch_date(
            mistral,
            cfg.brave_key,
            extracted,
            page_date,
            page_evidence,
            cfg.days_lookback,
        )
        if not date_result.get("within_window", False):
            logging.info(
                "Dropping %s: AI-validated launch date %s not in window",
                extracted.get("product_name", "?"),
                date_result.get("ai_launch_date"),
            )
            continue

        if cfg.skip_distributor:
            distributor_result = {"leads": [], "summary": "Distributor lookup skipped"}
        else:
            distributor_result = lookup_distributors(mistral, cfg.brave_key, extracted)

        rows.append(build_row(cand, extracted, date_result, distributor_result))
    return rows


def scan_web(cfg: ScannerConfig, mistral: MistralClient) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for source in WEB_SOURCES:
        logging.info("Web source: %s", source["name"])
        try:
            cands = build_web_candidates(cfg.brave_key, source, cfg.per_source)
            source_rows = process_candidates(cfg, mistral, cands, fetch_page=True)
        except Exception as exc:  # noqa: BLE001
            logging.exception("Web source %s failed: %s", source["name"], exc)
            source_rows = []
        if source_rows:
            rows.extend(source_rows)
        else:
            rows.append(
                empty_row("Web", source["name"], source["market"], "No qualifying product in window")
            )
    return rows


def scan_youtube(cfg: ScannerConfig, mistral: MistralClient) -> List[Dict[str, Any]]:
    logging.info("YouTube scan")
    try:
        cands = youtube_candidates(cfg.youtube_key, cfg.days_lookback)
        rows = process_candidates(cfg, mistral, cands, fetch_page=False)
    except Exception as exc:  # noqa: BLE001
        logging.exception("YouTube scan failed: %s", exc)
        rows = []
    if not rows:
        rows.append(empty_row("YouTube", "YouTube", "International", "No qualifying product in window"))
    return rows


def scan_social(cfg: ScannerConfig, mistral: MistralClient) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for platform, actor_id in APIFY_ACTORS.items():
        logging.info("Social source: %s (actor=%s)", platform, actor_id)
        try:
            items = run_apify_actor(
                cfg.apify_token, platform, actor_id, cfg.per_source * 4
            )
            cands = apify_items_to_candidates(platform, items)
            platform_rows = process_candidates(cfg, mistral, cands, fetch_page=False)
        except Exception as exc:  # noqa: BLE001
            logging.exception("Social %s failed: %s", platform, exc)
            platform_rows = []
        if platform_rows:
            rows.extend(platform_rows)
        else:
            rows.append(
                empty_row("Social", platform, social_market_for(platform), "No qualifying product in window")
            )
    return rows


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------


def download_image(image_url: str, output_dir: Path) -> Optional[Path]:
    if not image_url:
        return None
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        )
    }
    try:
        with requests.Session() as session:
            session.trust_env = not IGNORE_SYSTEM_PROXY
            r = session.get(image_url, headers=headers, timeout=REQUEST_TIMEOUT_SECONDS, stream=True)
        r.raise_for_status()
        ct = r.headers.get("Content-Type", "").lower()
        if "image" not in ct and not image_url.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
            return None
        raw = output_dir / safe_filename(image_url, ".raw")
        with raw.open("wb") as fh:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    fh.write(chunk)
        with Image.open(raw) as img:
            img = ImageOps.exif_transpose(img)
            img.thumbnail((120, 120), Image.Resampling.LANCZOS)
            canvas = Image.new("RGB", (120, 120), "white")
            ox = (120 - img.width) // 2
            oy = (120 - img.height) // 2
            if img.mode in ("RGBA", "LA"):
                canvas.paste(img, (ox, oy), img.convert("RGBA"))
            else:
                canvas.paste(img.convert("RGB"), (ox, oy))
            final = output_dir / safe_filename(image_url, ".png")
            canvas.save(final, format="PNG")
        try:
            raw.unlink(missing_ok=True)
        except OSError:
            pass
        return final
    except (requests.RequestException, UnidentifiedImageError, OSError) as exc:
        logging.info("Image download failed for %s: %s", image_url, exc)
        return None


def write_excel(rows: List[Dict[str, Any]], output_file: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Beauty Device Scan"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, col in enumerate(REQUIRED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [14, 22, 24, 22, 32, 18, 18, 44, 22, 38, 16, 30, 30, 38, 60, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REQUIRED_COLUMNS))}1"

    with tempfile.TemporaryDirectory(prefix="bd_imgs_") as tmpdir:
        tmp_path = Path(tmpdir)
        for r_idx, row in enumerate(rows, start=2):
            ws.row_dimensions[r_idx].height = 95
            for c_idx, col in enumerate(REQUIRED_COLUMNS, start=1):
                value = row.get(col, "N/A")
                if col == "Product Image":
                    value = "" if row.get("_image_url") else "No image"
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            image_url = row.get("_image_url") or ""
            if image_url:
                ipath = download_image(image_url, tmp_path)
                if ipath:
                    excel_img = ExcelImage(str(ipath))
                    excel_img.width = 120
                    excel_img.height = 120
                    col_letter = get_column_letter(REQUIRED_COLUMNS.index("Product Image") + 1)
                    ws.add_image(excel_img, f"{col_letter}{r_idx}")

        wb.save(output_file)
    logging.info("Saved Excel: %s", output_file)


# ---------------------------------------------------------------------------
# DRY-RUN FIXTURE
# ---------------------------------------------------------------------------


def dry_run_rows() -> List[Dict[str, Any]]:
    """Return a small fixture of fully-populated rows for offline smoke tests."""
    return [
        {
            "Source Type": "Social",
            "Source Name": "Xiaohongshu",
            "Country / Market": "Mainland China",
            "Product Category": "LED Face Mask",
            "Product Name": "AURORA Glow Mask Pro X",
            "Brand": "AURORA",
            "Product Image": "",
            "Key Functions / Selling Points": "Multi-wavelength LED + NIR, 9 modes, neck attachment",
            "Launch Date (AI Verified)": "2026-03-15",
            "Date Source / Evidence": "[High] Brand newsroom post dated Mar 15 confirmed by 2 XHS KOL reviews",
            "Price": "RMB 2,980",
            "Supplier / Brand Information": "AURORA Beauty Tech Co., Ltd (Shenzhen)",
            "Engagement / Popularity Signal": "Views: 1.2M; Likes: 86k; Saves: 14k",
            "Relevance to HK Sourcing": "High - K-beauty parity at sub-3000 RMB, fits HK price band",
            "Distributor / Reseller Leads": "- AURORA HK Store (Official, HK): https://aurora.hk\n- BeautyAsia Trading (Distributor, HK): https://beautyasia.example",
            "Source URL": "https://www.xiaohongshu.com/explore/sample",
            "_image_url": "",
        },
        {
            "Source Type": "Web",
            "Source Name": "Amazon Japan",
            "Country / Market": "Japan",
            "Product Category": "RF Facial Device",
            "Product Name": "Yaman Photo Plus Prestige III",
            "Brand": "Yaman",
            "Product Image": "",
            "Key Functions / Selling Points": "RF + EMS + LED, redesigned head, IP rating",
            "Launch Date (AI Verified)": "2026-02-20",
            "Date Source / Evidence": "[Medium] @cosme review + press release page age",
            "Price": "JPY 79,200",
            "Supplier / Brand Information": "Yaman Co., Ltd",
            "Engagement / Popularity Signal": "Top 5 in 美容家電 ranking",
            "Relevance to HK Sourcing": "High - established brand with HK demand",
            "Distributor / Reseller Leads": "- Yaman Official HK: https://yaman.com.hk\n- Citysuper HK (Reseller, HK): https://citysuper.example",
            "Source URL": "https://www.amazon.co.jp/sample",
            "_image_url": "",
        },
    ]


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Foreign-market beauty-device scanner for HK sourcing")
    p.add_argument("--per-source", type=int, default=DEFAULT_PER_SOURCE,
                   help="Number of products per source/channel (trial run: 3-5).")
    p.add_argument("--days-lookback", type=int, default=DEFAULT_DAYS_LOOKBACK,
                   help="Lookback window in days for considering a product 'new'.")
    p.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="Output Excel filename.")
    p.add_argument(
        "--channels",
        default="social,web,youtube",
        help="Comma-separated channels: any of social, web, youtube.",
    )
    p.add_argument("--skip-distributor", action="store_true",
                   help="Skip the Brave+Mistral distributor lookup step.")
    p.add_argument("--dry-run", action="store_true",
                   help="Skip all network calls and write a fixture Excel for smoke testing.")
    p.add_argument("-v", "--verbose", action="store_true")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    setup_logging(args.verbose)

    if not 1 <= args.per_source <= 20:
        logging.error("--per-source must be between 1 and 20 (recommended 3-5)")
        return 2

    cfg = ScannerConfig(
        brave_key=DEFAULT_BRAVE_API_KEY,
        mistral_key=DEFAULT_MISTRAL_API_KEY,
        youtube_key=DEFAULT_YOUTUBE_API_KEY,
        apify_token=DEFAULT_APIFY_API_TOKEN,
        per_source=args.per_source,
        days_lookback=args.days_lookback,
        output_file=args.output,
        enabled_channels=[c.strip().lower() for c in args.channels.split(",") if c.strip()],
        dry_run=args.dry_run,
        skip_distributor=args.skip_distributor,
    )

    if cfg.dry_run:
        logging.info("Dry run: writing fixture workbook")
        write_excel(dry_run_rows(), cfg.output_file)
        return 0

    mistral = MistralClient(cfg.mistral_key)
    if not mistral.preflight():
        logging.error("Mistral unavailable: %s", mistral.failure_reason)
        return 3

    rows: List[Dict[str, Any]] = []
    if "social" in cfg.enabled_channels:
        rows.extend(scan_social(cfg, mistral))
    if "web" in cfg.enabled_channels:
        rows.extend(scan_web(cfg, mistral))
    if "youtube" in cfg.enabled_channels:
        rows.extend(scan_youtube(cfg, mistral))

    if not rows:
        rows.append(empty_row("N/A", "N/A", "N/A", "No channels produced data"))

    write_excel(rows, cfg.output_file)
    logging.info("Done. Rows: %s", len(rows))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
