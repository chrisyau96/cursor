from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote

import requests
from openpyxl import Workbook


APIFY_BASE_URL = "https://api.apify.com/v2"
MISTRAL_BASE_URL = "https://api.mistral.ai/v1"
BRAVE_SEARCH_URL = "https://api.search.brave.com/res/v1/web/search"

DEFAULT_OUTPUT_PREFIX = "beauty_device_sourcing_trial"
REQUEST_TIMEOUT_SECONDS = 35

TEXT_KEYS = [
    "title",
    "caption",
    "description",
    "text",
    "desc",
    "content",
    "videoDescription",
]
URL_KEYS = ["url", "postUrl", "webVideoUrl", "videoUrl", "link", "shortCodeUrl"]
DATE_KEYS = ["date", "createdAt", "timestamp", "publishedAt", "takenAt", "createTime", "uploadDate"]
AUTHOR_KEYS = ["author", "authorName", "username", "ownerUsername", "channelName"]

ENGAGEMENT_KEYS = [
    "likes",
    "likeCount",
    "diggCount",
    "comments",
    "commentCount",
    "shares",
    "shareCount",
    "views",
    "viewCount",
    "playCount",
    "collectCount",
    "saveCount",
]

OUTPUT_COLUMNS = [
    "source_platform",
    "market",
    "product_name",
    "brand",
    "category",
    "why_trendy",
    "validated_source_date",
    "date_validation_method",
    "date_confidence",
    "popularity_signal",
    "target_market",
    "sourcing_fit_reason",
    "source_url",
    "distributor_sites",
]


@dataclass(frozen=True)
class SourceConfig:
    platform: str
    market: str
    actor_env: str
    default_queries: list[str]


@dataclass
class SocialCandidate:
    source_platform: str
    market: str
    url: str
    text: str
    posted_at_raw: str
    engagement_score: int
    engagement_text: str
    author: str
    raw_payload: dict[str, Any]


SOURCES: list[SourceConfig] = [
    SourceConfig(
        platform="TikTok",
        market="International",
        actor_env="APIFY_ACTOR_TIKTOK_ID",
        default_queries=["beauty device", "led mask", "rf facial device", "ipl device"],
    ),
]


def configure_logging(level: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def request_json(
    method: str,
    url: str,
    *,
    headers: Optional[dict[str, str]] = None,
    params: Optional[dict[str, Any]] = None,
    json_body: Optional[dict[str, Any]] = None,
    retries: int = 2,
    timeout: int = REQUEST_TIMEOUT_SECONDS,
) -> Any:
    for attempt in range(retries + 1):
        try:
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                params=params,
                json=json_body,
                timeout=timeout,
            )
            if response.status_code in {429, 500, 502, 503, 504} and attempt < retries:
                wait_seconds = 2 + attempt * 4
                logging.warning(
                    "Retrying %s %s after HTTP %s (%ss wait)",
                    method,
                    url,
                    response.status_code,
                    wait_seconds,
                )
                time.sleep(wait_seconds)
                continue
            response.raise_for_status()
            if not response.text.strip():
                return {}
            return response.json()
        except (requests.RequestException, json.JSONDecodeError) as exc:
            if attempt == retries:
                raise RuntimeError(f"Request failed ({method} {url}): {exc}") from exc
            time.sleep(2 + attempt * 4)
    raise RuntimeError(f"Request failed ({method} {url})")


def clean_text(value: str, max_len: int = 1800) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:max_len]


def numeric_value(value: Any) -> int:
    if value is None:
        return 0
    try:
        return int(str(value).replace(",", "").strip())
    except ValueError:
        return 0


def pick_first_string(payload: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def extract_candidate(config: SourceConfig, item: dict[str, Any]) -> Optional[SocialCandidate]:
    text = " ".join(clean_text(item.get(key, "")) for key in TEXT_KEYS if item.get(key))
    if not text:
        text = clean_text(str(item))
    url = pick_first_string(item, URL_KEYS)
    posted_at_raw = pick_first_string(item, DATE_KEYS)
    author = pick_first_string(item, AUTHOR_KEYS)

    engagement_parts: list[str] = []
    engagement_score = 0
    for key in ENGAGEMENT_KEYS:
        if key in item:
            val = numeric_value(item.get(key))
            if val > 0:
                engagement_parts.append(f"{key}:{val}")
                engagement_score += val
    engagement_text = "; ".join(engagement_parts) if engagement_parts else "N/A"

    if not text and not url:
        return None

    return SocialCandidate(
        source_platform=config.platform,
        market=config.market,
        url=url or "N/A",
        text=text,
        posted_at_raw=posted_at_raw,
        engagement_score=engagement_score,
        engagement_text=engagement_text,
        author=author or "unknown",
        raw_payload=item,
    )


def run_apify_actor(
    *,
    token: str,
    actor_id: str,
    source: SourceConfig,
    max_input_items: int,
    custom_queries: Optional[list[str]] = None,
) -> list[dict[str, Any]]:
    actor_input = {
        "searchTerms": custom_queries or source.default_queries,
        "queries": custom_queries or source.default_queries,
        "maxItems": max_input_items,
        "resultsLimit": max_input_items,
        "sort": "recent",
    }
    encoded_actor = quote(actor_id, safe="")
    run_url = f"{APIFY_BASE_URL}/acts/{encoded_actor}/runs"
    run_data = request_json(
        "POST",
        run_url,
        params={"token": token, "waitForFinish": 180},
        json_body=actor_input,
        retries=1,
        timeout=190,
    )

    run_obj = run_data.get("data", run_data)
    status = run_obj.get("status", "")
    run_id = run_obj.get("id", "")
    dataset_id = run_obj.get("defaultDatasetId", "")

    if status not in {"SUCCEEDED", "READY"} and run_id:
        status_url = f"{APIFY_BASE_URL}/actor-runs/{run_id}"
        for _ in range(12):
            status_resp = request_json("GET", status_url, params={"token": token}, retries=0)
            run_obj = status_resp.get("data", {})
            status = run_obj.get("status", "")
            dataset_id = run_obj.get("defaultDatasetId", dataset_id)
            if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
                break
            time.sleep(8)

    if status != "SUCCEEDED" or not dataset_id:
        logging.warning("Apify actor %s for %s did not succeed (status=%s)", actor_id, source.platform, status)
        return []

    items_url = f"{APIFY_BASE_URL}/datasets/{dataset_id}/items"
    items = request_json(
        "GET",
        items_url,
        params={
            "token": token,
            "clean": "true",
            "format": "json",
            "limit": max_input_items,
        },
        retries=1,
    )
    return items if isinstance(items, list) else []


def parse_json_from_text(text: str) -> Any:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?", "", stripped, flags=re.IGNORECASE).strip()
        stripped = re.sub(r"```$", "", stripped).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        pass
    match = re.search(r"(\{.*\}|\[.*\])", stripped, flags=re.DOTALL)
    if not match:
        raise ValueError("No JSON content found in model response.")
    return json.loads(match.group(1))


def mistral_extract_products(
    *,
    api_key: str,
    model: str,
    source: SourceConfig,
    candidates: list[SocialCandidate],
    min_results: int,
    max_results: int,
    days_lookback: int,
) -> list[dict[str, Any]]:
    payload_candidates = [
        {
            "index": idx,
            "source_platform": c.source_platform,
            "market": c.market,
            "source_url": c.url,
            "post_text": c.text,
            "posted_at_raw": c.posted_at_raw,
            "engagement_signal": c.engagement_text,
            "author": c.author,
        }
        for idx, c in enumerate(candidates, start=1)
    ]
    user_prompt = f"""
Today is {dt.date.today().isoformat()}.
You are a product sourcing analyst focused on home-use beauty devices for Hong Kong.

Task:
1) From the candidates, select trendy NEW beauty-device products from foreign markets.
2) Return between {min_results} and {max_results} validated products for this source.
3) Validate source date by combining platform date metadata + post text clues + relative timing words (do not rely on one field only).
4) Exclude non-device cosmetics, old products outside ~{days_lookback} days, medical/clinic-only devices, or vague category-only mentions.

Output strict JSON with this schema:
{{
  "items": [
    {{
      "source_platform": "{source.platform}",
      "market": "{source.market}",
      "product_name": "...",
      "brand": "...",
      "category": "...",
      "why_trendy": "...",
      "validated_source_date": "YYYY-MM-DD or Unknown",
      "date_validation_method": "how date was validated",
      "date_confidence": "high|medium|low",
      "popularity_signal": "...",
      "target_market": "...",
      "sourcing_fit_reason": "...",
      "source_url": "..."
    }}
  ]
}}

Candidates:
{json.dumps(payload_candidates, ensure_ascii=False)}
""".strip()
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": "Return one JSON object only."},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.1,
        "max_tokens": 1800,
        "response_format": {"type": "json_object"},
    }
    response = request_json(
        "POST",
        f"{MISTRAL_BASE_URL}/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json_body=body,
        retries=1,
    )
    content = response["choices"][0]["message"]["content"]
    data = parse_json_from_text(content)
    items = data.get("items", []) if isinstance(data, dict) else []
    if not isinstance(items, list):
        return []
    normalized = [normalize_model_item(item, source) for item in items]
    normalized = [item for item in normalized if item.get("product_name") not in {"", "Unknown", "N/A"}]
    return normalized[:max_results]


def normalize_model_item(item: dict[str, Any], source: SourceConfig) -> dict[str, Any]:
    def get(key: str, fallback: str = "") -> str:
        value = item.get(key, fallback)
        return clean_text(value, 2000)

    return {
        "source_platform": get("source_platform") or source.platform,
        "market": get("market") or source.market,
        "product_name": get("product_name") or "Unknown",
        "brand": get("brand") or "Unknown",
        "category": get("category") or "Beauty device",
        "why_trendy": get("why_trendy") or "N/A",
        "validated_source_date": get("validated_source_date") or "Unknown",
        "date_validation_method": get("date_validation_method") or "N/A",
        "date_confidence": (get("date_confidence") or "low").lower(),
        "popularity_signal": get("popularity_signal") or "N/A",
        "target_market": get("target_market") or source.market,
        "sourcing_fit_reason": get("sourcing_fit_reason") or "N/A",
        "source_url": get("source_url") or "N/A",
        "distributor_sites": "",
    }


def heuristic_extract_products(
    source: SourceConfig,
    candidates: list[SocialCandidate],
    *,
    max_results: int,
) -> list[dict[str, Any]]:
    keywords = ["led", "microcurrent", "rf", "ipl", "beauty device", "美容仪", "美顔器", "skin device"]
    results: list[dict[str, Any]] = []
    for candidate in candidates:
        lowered = candidate.text.lower()
        if not any(term in lowered for term in keywords):
            continue
        title_guess = clean_text(candidate.text, 120).split(".")[0]
        words = title_guess.split(" ")
        product_name = " ".join(words[:8]) if words else "Unknown product"
        results.append(
            {
                "source_platform": source.platform,
                "market": source.market,
                "product_name": product_name,
                "brand": "Unknown",
                "category": "Beauty device",
                "why_trendy": "High social engagement and recurring beauty-device keywords.",
                "validated_source_date": parse_candidate_date(candidate.posted_at_raw),
                "date_validation_method": "Heuristic fallback from post metadata; enable Mistral for AI validation.",
                "date_confidence": "low",
                "popularity_signal": candidate.engagement_text,
                "target_market": source.market,
                "sourcing_fit_reason": "Needs AI validation before supplier decision.",
                "source_url": candidate.url,
                "distributor_sites": "",
            }
        )
        if len(results) >= max_results:
            break
    return results


def parse_candidate_date(raw_value: str) -> str:
    if not raw_value:
        return dt.date.today().isoformat()
    cleaned = str(raw_value).strip()
    try:
        return dt.datetime.fromisoformat(cleaned.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        pass
    match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", cleaned)
    if match:
        y, m, d = match.groups()
        return dt.date(int(y), int(m), int(d)).isoformat()
    return dt.date.today().isoformat()


def brave_search(query: str, api_key: str, count: int = 5) -> list[dict[str, Any]]:
    data = request_json(
        "GET",
        BRAVE_SEARCH_URL,
        headers={"Accept": "application/json", "X-Subscription-Token": api_key},
        params={
            "q": query,
            "count": min(20, count),
            "result_filter": "web",
            "text_decorations": "false",
        },
        retries=1,
    )
    return data.get("web", {}).get("results", []) if isinstance(data, dict) else []


def rank_distributor_sites(search_results: list[dict[str, Any]]) -> list[str]:
    scored: list[tuple[int, str]] = []
    for result in search_results:
        url = result.get("url", "")
        title = clean_text(result.get("title", ""))
        description = clean_text(result.get("description", ""))
        combined = f"{title} {description} {url}".lower()
        if not url.startswith("http"):
            continue
        score = 0
        if "distributor" in combined:
            score += 4
        if "official" in combined or "authorized" in combined:
            score += 3
        if ".hk" in url or "hong kong" in combined:
            score += 3
        if "store" in combined or "shop" in combined:
            score += 1
        scored.append((score, url))

    scored.sort(key=lambda pair: pair[0], reverse=True)
    unique_urls: list[str] = []
    for _, url in scored:
        if url not in unique_urls:
            unique_urls.append(url)
        if len(unique_urls) >= 3:
            break
    return unique_urls


def enrich_with_distributors(items: list[dict[str, Any]], brave_api_key: str) -> None:
    for item in items:
        query = (
            f"{item.get('brand', '')} {item.get('product_name', '')} "
            "official distributor Hong Kong"
        ).strip()
        if not query:
            item["distributor_sites"] = ""
            continue
        try:
            search_results = brave_search(query, brave_api_key, count=7)
            urls = rank_distributor_sites(search_results)
            item["distributor_sites"] = "; ".join(urls)
            time.sleep(1.2)
        except RuntimeError as exc:
            logging.warning("Distributor lookup failed for %s: %s", item.get("product_name", "N/A"), exc)
            item["distributor_sites"] = ""


def deduplicate(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    unique_items: list[dict[str, Any]] = []
    for item in items:
        key = f"{item.get('source_platform', '').lower()}|{item.get('product_name', '').lower()}|{item.get('brand', '').lower()}"
        if key in seen:
            continue
        seen.add(key)
        unique_items.append(item)
    return unique_items


def write_outputs(items: list[dict[str, Any]], output_prefix: str) -> tuple[Path, Path]:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = Path(f"{output_prefix}_{timestamp}.json")
    xlsx_path = Path(f"{output_prefix}_{timestamp}.xlsx")

    with json_path.open("w", encoding="utf-8") as fp:
        json.dump(items, fp, ensure_ascii=False, indent=2)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Beauty Device Trends"
    for col_idx, column in enumerate(OUTPUT_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=column)

    for row_idx, item in enumerate(items, start=2):
        for col_idx, column in enumerate(OUTPUT_COLUMNS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=item.get(column, ""))

    workbook.save(xlsx_path)
    return json_path, xlsx_path


def load_fixture(path: Path) -> dict[str, list[dict[str, Any]]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Offline fixture must be a JSON object keyed by source platform.")
    normalized: dict[str, list[dict[str, Any]]] = {}
    for platform, items in data.items():
        normalized[str(platform)] = items if isinstance(items, list) else []
    return normalized


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Discover trendy foreign beauty devices for Hong Kong sourcing via Apify + AI validation."
    )
    parser.add_argument("--min-results-per-source", type=int, default=3)
    parser.add_argument("--max-results-per-source", type=int, default=5)
    parser.add_argument("--max-input-items-per-source", type=int, default=20)
    parser.add_argument("--days-lookback", type=int, default=120)
    parser.add_argument("--output-prefix", type=str, default=DEFAULT_OUTPUT_PREFIX)
    parser.add_argument("--mistral-model", type=str, default="mistral-small-latest")
    parser.add_argument("--offline-fixture", type=str, default="")
    parser.add_argument("--skip-distributor-search", action="store_true")
    parser.add_argument("--log-level", type=str, default="INFO")
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.min_results_per_source < 3 or args.max_results_per_source > 5:
        raise ValueError("Trial mode requires 3-5 results per source. Keep min>=3 and max<=5.")
    if args.min_results_per_source > args.max_results_per_source:
        raise ValueError("min-results-per-source cannot exceed max-results-per-source.")
    if args.max_input_items_per_source < args.max_results_per_source:
        raise ValueError("max-input-items-per-source must be >= max-results-per-source.")


def collect_source_candidates(
    source: SourceConfig,
    *,
    apify_token: str,
    max_input_items: int,
    fixture_data: Optional[dict[str, list[dict[str, Any]]]],
) -> list[SocialCandidate]:
    raw_items: list[dict[str, Any]] = []
    if fixture_data is not None:
        raw_items = fixture_data.get(source.platform, [])
    else:
        actor_id = os.getenv(source.actor_env, "").strip()
        if not actor_id:
            logging.warning("Skipping %s because %s is not configured.", source.platform, source.actor_env)
            return []
        raw_items = run_apify_actor(
            token=apify_token,
            actor_id=actor_id,
            source=source,
            max_input_items=max_input_items,
        )

    candidates: list[SocialCandidate] = []
    for item in raw_items:
        candidate = extract_candidate(source, item)
        if candidate:
            candidates.append(candidate)

    candidates.sort(key=lambda c: c.engagement_score, reverse=True)
    return candidates[:max_input_items]


def process_source(
    source: SourceConfig,
    *,
    candidates: list[SocialCandidate],
    mistral_api_key: str,
    mistral_model: str,
    min_results: int,
    max_results: int,
    days_lookback: int,
) -> list[dict[str, Any]]:
    if not candidates:
        return []
    extracted_items: list[dict[str, Any]] = []
    if mistral_api_key:
        try:
            extracted_items = mistral_extract_products(
                api_key=mistral_api_key,
                model=mistral_model,
                source=source,
                candidates=candidates,
                min_results=min_results,
                max_results=max_results,
                days_lookback=days_lookback,
            )
        except (RuntimeError, KeyError, ValueError) as exc:
            logging.warning("Mistral extraction failed for %s: %s", source.platform, exc)
            extracted_items = []

    if not extracted_items:
        logging.warning("Using heuristic fallback for %s (configure MISTRAL_API_KEY for AI validation).", source.platform)
        extracted_items = heuristic_extract_products(source, candidates, max_results=max_results)

    if len(extracted_items) < min_results:
        extracted_items = backfill_to_min_results(
            source,
            candidates=candidates,
            existing=extracted_items,
            min_results=min_results,
            max_results=max_results,
        )
    return extracted_items[:max_results]


def backfill_to_min_results(
    source: SourceConfig,
    *,
    candidates: list[SocialCandidate],
    existing: list[dict[str, Any]],
    min_results: int,
    max_results: int,
) -> list[dict[str, Any]]:
    if len(existing) >= min_results:
        return existing

    seen_urls = {item.get("source_url", "") for item in existing}
    for candidate in candidates:
        if candidate.url in seen_urls:
            continue
        existing.append(
            {
                "source_platform": source.platform,
                "market": source.market,
                "product_name": clean_text(candidate.text, 120).split(".")[0] or "Unknown product",
                "brand": "Unknown",
                "category": "Beauty device",
                "why_trendy": "Backfilled trial candidate to satisfy 3-5 sourcing trial volume.",
                "validated_source_date": parse_candidate_date(candidate.posted_at_raw),
                "date_validation_method": "Backfill from post metadata; replace with AI-validated date in production run.",
                "date_confidence": "low",
                "popularity_signal": candidate.engagement_text,
                "target_market": source.market,
                "sourcing_fit_reason": "Trial candidate requiring AI/date and supplier confirmation.",
                "source_url": candidate.url,
                "distributor_sites": "",
            }
        )
        seen_urls.add(candidate.url)
        if len(existing) >= min_results or len(existing) >= max_results:
            break
    return existing


def main() -> None:
    args = parse_args()
    configure_logging(args.log_level)
    validate_args(args)

    apify_token = os.getenv("APIFY_API_TOKEN", "").strip()
    mistral_api_key = os.getenv("MISTRAL_API_KEY", "").strip()
    brave_api_key = os.getenv("BRAVE_API_KEY", "").strip()
    fixture_data: Optional[dict[str, list[dict[str, Any]]]] = None

    if args.offline_fixture:
        fixture_data = load_fixture(Path(args.offline_fixture))
        logging.info("Loaded offline fixture from %s", args.offline_fixture)
    elif not apify_token:
        raise ValueError("APIFY_API_TOKEN is required unless --offline-fixture is provided.")

    all_items: list[dict[str, Any]] = []
    for source in SOURCES:
        logging.info("Collecting candidates from %s...", source.platform)
        candidates = collect_source_candidates(
            source,
            apify_token=apify_token,
            max_input_items=args.max_input_items_per_source,
            fixture_data=fixture_data,
        )
        source_items = process_source(
            source,
            candidates=candidates,
            mistral_api_key=mistral_api_key,
            mistral_model=args.mistral_model,
            min_results=args.min_results_per_source,
            max_results=args.max_results_per_source,
            days_lookback=args.days_lookback,
        )
        if len(source_items) < args.min_results_per_source:
            logging.warning(
                "%s returned %s items (below target min=%s).",
                source.platform,
                len(source_items),
                args.min_results_per_source,
            )
        all_items.extend(source_items[: args.max_results_per_source])
        time.sleep(1)

    all_items = deduplicate(all_items)

    if all_items and not args.skip_distributor_search and brave_api_key:
        logging.info("Running distributor-site lookup...")
        enrich_with_distributors(all_items, brave_api_key)
    elif not brave_api_key and not args.skip_distributor_search:
        logging.warning("BRAVE_API_KEY is not configured; skipping distributor-site lookup.")

    json_path, xlsx_path = write_outputs(all_items, args.output_prefix)
    logging.info("Completed. Exported %s products.", len(all_items))
    logging.info("JSON output: %s", json_path)
    logging.info("Excel output: %s", xlsx_path)


if __name__ == "__main__":
    main()
